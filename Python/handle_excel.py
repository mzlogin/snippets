"""
使用场景：
有三个巨大的 Excel 文件（合计约一百三十万行）存放在 src_new 下面，
这个脚本能读取这三个 Excel 文件，将所有数据按第 5 列的值分别存放到 dst_new 里的不同的文件里去，
dst 文件以数据第 5 列值命名
template_new.xlsx 是一个模板文件，只含表头

Python 3
"""

import openpyxl
import os
import shutil

base_path = '/Users/mazhuang/Downloads/excels/'

src_path = base_path + 'src_new/'

dst_path = base_path + 'dst_new/'

template_file = base_path + 'template_new.xlsx'

memory_dict = {}

total_rows = 0


def read_from_file(file):

    global total_rows
    global memory_dict

    src_workbook = openpyxl.load_workbook(filename=file, read_only=True)
    sheet_names = src_workbook.sheetnames
    for sheet_name in sheet_names:
        src_sheet = src_workbook[sheet_name]

        count = 0
        for row in src_sheet.rows:

            count += 1
            if count == 1:
                continue

            total_rows += 1

            sku_group = row[4].value

            if sku_group == '':
                print('存在数据物料组为空')
                sku_group = '物料组为空'

            if sku_group not in memory_dict:
                memory_dict[sku_group] = []

            memory_dict[sku_group].append([cell.value for cell in row])


for file in os.listdir(src_path):
    if file.endswith('xlsx'):
        read_from_file(src_path + file)

print('total rows: %d' % total_rows)

dst_rows = 0

for key, value in memory_dict.items():

    dst_rows += len(value)

    print('%s, %d' % (key, len(value)))

    dst_file = dst_path + key + '.xlsx'
    if not os.path.exists(dst_file):
        shutil.copy(template_file, dst_file)

    dst_workbook = openpyxl.load_workbook(filename=dst_file)
    for dst_sheet_name in dst_workbook.sheetnames:
        dst_sheet = dst_workbook[dst_sheet_name]
        for row in value:
            dst_sheet.append(row)
    dst_workbook.save(dst_file)

print('sku groups: %d' % len(memory_dict))

print('dst rows: %d' % dst_rows)
